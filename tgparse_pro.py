"""
TGParse PRO — Коммерческий парсер с оплатой
============================================
Модель: оплата за кол-во сообщений (Telegram Stars)
Хостинг: Railway / Render

Тарифы:
  FREE  — до 100 сообщений, 1 чат
  BASIC — 1000 сообщений, 3 чата, 50 Stars
  PRO   — 10000 сообщений, 10 чатов + автопарсинг, 150 Stars
  MAX   — 50000 сообщений, безлимит чатов + расписание, 400 Stars

Установка:
    pip install python-telegram-bot telethon pandas openpyxl python-dotenv apscheduler

Запуск:
    python tgparse_pro.py
"""

import asyncio
import io
import logging
import os
import sqlite3
from datetime import datetime, timezone, timedelta
from typing import Optional

import pandas as pd
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from dotenv import load_dotenv
from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    LabeledPrice, ReplyKeyboardMarkup, KeyboardButton,
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    ConversationHandler, PreCheckoutQueryHandler, ContextTypes, filters,
)
from telethon import TelegramClient
from telethon.tl.types import User

load_dotenv()

BOT_TOKEN = os.getenv("TG_BOT_TOKEN")
API_ID    = int(os.getenv("TG_API_ID", "0"))
API_HASH  = os.getenv("TG_API_HASH", "")
SESSION   = os.getenv("TG_SESSION", "tgparse")
ADMIN_ID  = int(os.getenv("ADMIN_ID", "0"))  # Твой Telegram ID для статистики

logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO)
log = logging.getLogger(__name__)

# ─── Тарифы ──────────────────────────────────────────────────────────────────
PLANS = {
    "free": {
        "name": "🆓 FREE",
        "price": 0,
        "msg_limit": 100,
        "chat_limit": 1,
        "scheduler": False,
        "description": "100 сообщений, 1 чат",
    },
    "basic": {
        "name": "⚡ BASIC",
        "price": 50,         # Telegram Stars
        "msg_limit": 1000,
        "chat_limit": 3,
        "scheduler": False,
        "description": "1 000 сообщений, 3 чата",
    },
    "pro": {
        "name": "🚀 PRO",
        "price": 150,
        "msg_limit": 10000,
        "chat_limit": 10,
        "scheduler": True,
        "description": "10 000 сообщений, 10 чатов, автопарсинг",
    },
    "max": {
        "name": "💎 MAX",
        "price": 400,
        "msg_limit": 50000,
        "chat_limit": 999,
        "scheduler": True,
        "description": "50 000 сообщений, ∞ чатов, расписание",
    },
}

# ─── Состояния диалога ────────────────────────────────────────────────────────
(WAIT_CHAT, WAIT_PERIOD, WAIT_LIMIT, WAIT_FORMAT,
 WAIT_SCHEDULE_CHAT, WAIT_SCHEDULE_INTERVAL) = range(6)

# ─── База данных ─────────────────────────────────────────────────────────────
DB = "tgparse.db"

def db_init():
    con = sqlite3.connect(DB)
    cur = con.cursor()
    cur.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            user_id     INTEGER PRIMARY KEY,
            username    TEXT,
            plan        TEXT DEFAULT 'free',
            plan_until  TEXT,
            msgs_used   INTEGER DEFAULT 0,
            created_at  TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS payments (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER,
            plan        TEXT,
            stars       INTEGER,
            created_at  TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS schedules (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER,
            chat        TEXT,
            interval_h  INTEGER,
            last_run    TEXT,
            active      INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS parse_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER,
            chat        TEXT,
            msg_count   INTEGER,
            created_at  TEXT DEFAULT CURRENT_TIMESTAMP
        );
    """)
    con.commit()
    con.close()

def db_get_user(user_id: int) -> dict:
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    row = con.execute("SELECT * FROM users WHERE user_id=?", (user_id,)).fetchone()
    con.close()
    if not row:
        return {"user_id": user_id, "plan": "free", "msgs_used": 0, "plan_until": None}
    return dict(row)

def db_upsert_user(user_id: int, username: str):
    con = sqlite3.connect(DB)
    con.execute("""
        INSERT INTO users(user_id, username) VALUES(?,?)
        ON CONFLICT(user_id) DO UPDATE SET username=excluded.username
    """, (user_id, username or ""))
    con.commit(); con.close()

def db_set_plan(user_id: int, plan: str, days: int = 30):
    until = (datetime.now() + timedelta(days=days)).isoformat()
    con = sqlite3.connect(DB)
    con.execute("UPDATE users SET plan=?, plan_until=? WHERE user_id=?", (plan, until, user_id))
    con.commit(); con.close()

def db_add_payment(user_id: int, plan: str, stars: int):
    con = sqlite3.connect(DB)
    con.execute("INSERT INTO payments(user_id,plan,stars) VALUES(?,?,?)", (user_id, plan, stars))
    con.commit(); con.close()

def db_log_parse(user_id: int, chat: str, count: int):
    con = sqlite3.connect(DB)
    con.execute("INSERT INTO parse_log(user_id,chat,msg_count) VALUES(?,?,?)", (user_id, chat, count))
    con.execute("UPDATE users SET msgs_used=msgs_used+? WHERE user_id=?", (count, user_id))
    con.commit(); con.close()

def db_get_schedules(active_only=True) -> list:
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    q = "SELECT * FROM schedules" + (" WHERE active=1" if active_only else "")
    rows = con.execute(q).fetchall()
    con.close()
    return [dict(r) for r in rows]

def db_add_schedule(user_id: int, chat: str, interval_h: int):
    con = sqlite3.connect(DB)
    con.execute("INSERT INTO schedules(user_id,chat,interval_h) VALUES(?,?,?)", (user_id, chat, interval_h))
    con.commit(); con.close()

def db_stats() -> dict:
    con = sqlite3.connect(DB)
    stats = {
        "users":    con.execute("SELECT COUNT(*) FROM users").fetchone()[0],
        "paid":     con.execute("SELECT COUNT(*) FROM users WHERE plan!='free'").fetchone()[0],
        "revenue":  con.execute("SELECT SUM(stars) FROM payments").fetchone()[0] or 0,
        "parses":   con.execute("SELECT COUNT(*) FROM parse_log").fetchone()[0],
        "msgs":     con.execute("SELECT SUM(msg_count) FROM parse_log").fetchone()[0] or 0,
    }
    con.close()
    return stats

# ─── Проверка плана ───────────────────────────────────────────────────────────
def get_user_plan(user_id: int) -> dict:
    u = db_get_user(user_id)
    plan_key = u.get("plan", "free")
    # Проверяем не истёк ли план
    if plan_key != "free" and u.get("plan_until"):
        if datetime.fromisoformat(u["plan_until"]) < datetime.now():
            plan_key = "free"
    return PLANS.get(plan_key, PLANS["free"])

# ─── Парсер ───────────────────────────────────────────────────────────────────
async def parse_messages(chat, date_from, date_to, limit, keywords=None, progress_cb=None):
    rows = []
    async with TelegramClient(SESSION, API_ID, API_HASH) as client:
        try:
            entity = await client.get_entity(chat)
        except Exception as e:
            raise ValueError(f"Чат не найден: {e}")

        chat_title = getattr(entity, "title", chat)
        iter_kw = dict(entity=entity, limit=limit)
        if date_to:   iter_kw["offset_date"] = date_to
        if date_from: iter_kw["reverse"] = True; iter_kw["offset_date"] = date_from

        count = 0
        async for msg in client.iter_messages(**iter_kw):
            msg_date = msg.date.replace(tzinfo=timezone.utc)
            if date_from and msg_date < date_from: continue
            if date_to   and msg_date > date_to:   continue
            if not msg.text: continue
            if keywords and not any(k.lower() in msg.text.lower() for k in keywords): continue

            try:
                sender = await msg.get_sender()
                if isinstance(sender, User):
                    uname = f"@{sender.username}" if sender.username else f"{sender.first_name or ''} {sender.last_name or ''}".strip()
                else:
                    uname = getattr(sender, "title", "Unknown")
            except Exception:
                uname = "Unknown"

            rows.append({
                "№": len(rows)+1,
                "Пользователь": uname,
                "Запрос": msg.text[:500],
                "Чат": chat_title,
                "Дата": msg_date.strftime("%d.%m.%Y %H:%M"),
            })
            count += 1
            if progress_cb and count % 100 == 0:
                await progress_cb(count)

    return pd.DataFrame(rows)

def df_to_excel(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Парсинг")
        ws = w.sheets["Парсинг"]
        from openpyxl.styles import PatternFill, Font, Alignment
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col)+4, 60)
    buf.seek(0); return buf.read()

def df_to_csv(df):
    return df.to_csv(index=False).encode("utf-8-sig")

# ─── /start ───────────────────────────────────────────────────────────────────
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    db_upsert_user(u.id, u.username)
    plan = get_user_plan(u.id)

    kb = ReplyKeyboardMarkup([
        [KeyboardButton("▶ Парсить"), KeyboardButton("💳 Тарифы")],
        [KeyboardButton("📅 Расписание"), KeyboardButton("📊 Мой аккаунт")],
    ], resize_keyboard=True)

    await update.message.reply_text(
        f"👋 Привет, {u.first_name}!\n\n"
        f"🤖 *TGParse PRO* — профессиональный парсер Telegram\n\n"
        f"Твой тариф: *{plan['name']}*\n"
        f"Лимит: *{plan['msg_limit']:,} сообщений*, *{plan['chat_limit']} чатов*\n\n"
        f"Выбери действие 👇",
        parse_mode="Markdown", reply_markup=kb,
    )

# ─── Тарифы ───────────────────────────────────────────────────────────────────
async def cmd_plans(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    user_plan = get_user_plan(update.effective_user.id)
    text = "💳 *Тарифы TGParse PRO*\n\n"
    kb = []

    for key, p in PLANS.items():
        current = "✅ " if p["name"] == user_plan["name"] else ""
        sched = "⏰ Автопарсинг " if p["scheduler"] else ""
        text += (
            f"{current}*{p['name']}* — {p['description']}\n"
            f"{sched}{'Бесплатно' if p['price']==0 else f'{p[chr(112)+(chr(114)+chr(105)+chr(99)+chr(101))]}'+'⭐'}\n\n"
        )
        if p["price"] > 0 and p["name"] != user_plan["name"]:
            kb.append([InlineKeyboardButton(
                f"Купить {p['name']} — {p['price']}⭐",
                callback_data=f"buy_{key}"
            )])

    await update.message.reply_text(
        text, parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(kb) if kb else None,
    )

# ─── Оплата (Telegram Stars) ─────────────────────────────────────────────────
async def handle_buy(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    plan_key = query.data.replace("buy_", "")
    plan = PLANS.get(plan_key)
    if not plan: return

    await ctx.bot.send_invoice(
        chat_id=query.from_user.id,
        title=f"TGParse {plan['name']}",
        description=f"{plan['description']} • 30 дней",
        payload=f"plan_{plan_key}",
        currency="XTR",                    # Telegram Stars
        prices=[LabeledPrice(plan["name"], plan["price"])],
        provider_token="",                 # Пустой для Stars
    )

async def pre_checkout(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.pre_checkout_query.answer(ok=True)

async def payment_success(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    payload = update.message.successful_payment.invoice_payload
    stars   = update.message.successful_payment.total_amount
    user_id = update.effective_user.id
    plan_key = payload.replace("plan_", "")
    plan = PLANS.get(plan_key)

    db_set_plan(user_id, plan_key, days=30)
    db_add_payment(user_id, plan_key, stars)

    await update.message.reply_text(
        f"✅ *Оплата прошла успешно!*\n\n"
        f"Тариф *{plan['name']}* активирован на 30 дней.\n"
        f"Лимит: {plan['msg_limit']:,} сообщений, {plan['chat_limit']} чатов\n\n"
        f"Используй /parse для начала парсинга 🚀",
        parse_mode="Markdown",
    )

# ─── Диалог /parse ────────────────────────────────────────────────────────────
async def cmd_parse(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    db_upsert_user(user_id, update.effective_user.username)
    plan = get_user_plan(user_id)
    ctx.user_data["plan"] = plan

    await update.message.reply_text(
        f"📡 *Парсинг чата*\n\n"
        f"Твой тариф: *{plan['name']}*\n"
        f"Макс. сообщений: *{plan['msg_limit']:,}*\n"
        f"Макс. чатов: *{plan['chat_limit']}*\n\n"
        f"Введи @username или ссылку на чат:",
        parse_mode="Markdown",
    )
    return WAIT_CHAT

async def got_chat(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chats_raw = update.message.text.strip()
    plan = ctx.user_data.get("plan", PLANS["free"])

    # Несколько чатов через запятую
    chats = [c.strip() for c in chats_raw.split(",")]
    if len(chats) > plan["chat_limit"]:
        await update.message.reply_text(
            f"⚠️ Твой тариф позволяет парсить не более *{plan['chat_limit']} чатов* за раз.\n"
            f"Ты ввёл {len(chats)}. Повысь тариф — /plans",
            parse_mode="Markdown",
        )
        return WAIT_CHAT

    ctx.user_data["chats"] = chats

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 Сегодня",   callback_data="p_today"),
         InlineKeyboardButton("📅 7 дней",    callback_data="p_7days")],
        [InlineKeyboardButton("📅 30 дней",   callback_data="p_30days"),
         InlineKeyboardButton("📅 90 дней",   callback_data="p_90days")],
        [InlineKeyboardButton("📅 Всё время", callback_data="p_all")],
    ])
    await update.message.reply_text("📅 Выбери период:", reply_markup=kb)
    return WAIT_PERIOD

async def got_period(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    period = q.data.replace("p_", "")
    ctx.user_data["period"] = period
    plan = ctx.user_data.get("plan", PLANS["free"])

    limits = [100, 500, 1000, 5000, 10000, 50000]
    available = [l for l in limits if l <= plan["msg_limit"]]

    rows = []
    for i in range(0, len(available), 2):
        row = [InlineKeyboardButton(f"{available[i]:,}", callback_data=f"l_{available[i]}")]
        if i+1 < len(available):
            row.append(InlineKeyboardButton(f"{available[i+1]:,}", callback_data=f"l_{available[i+1]}"))
        rows.append(row)

    await q.edit_message_text(
        f"✅ Период выбран\n\n🔢 Сколько сообщений собрать?\n_(макс. {plan['msg_limit']:,} на твоём тарифе)_",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(rows),
    )
    return WAIT_LIMIT

async def got_limit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    ctx.user_data["limit"] = int(q.data.replace("l_", ""))

    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("📊 Excel", callback_data="f_excel"),
        InlineKeyboardButton("📄 CSV",   callback_data="f_csv"),
    ]])
    await q.edit_message_text("💾 Выбери формат файла:", reply_markup=kb)
    return WAIT_FORMAT

async def got_format(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    fmt     = q.data.replace("f_", "")
    chats   = ctx.user_data["chats"]
    period  = ctx.user_data["period"]
    limit   = ctx.user_data["limit"]
    user_id = q.from_user.id

    periods_map = {
        "today":  (datetime.now(timezone.utc).replace(hour=0,minute=0,second=0), None),
        "7days":  (datetime.now(timezone.utc)-timedelta(days=7), None),
        "30days": (datetime.now(timezone.utc)-timedelta(days=30), None),
        "90days": (datetime.now(timezone.utc)-timedelta(days=90), None),
        "all":    (None, None),
    }
    date_from, date_to = periods_map.get(period, (None, None))

    await q.edit_message_text(
        f"⚙️ Запускаю парсинг {len(chats)} чат(ов)...\n⏳ Подожди немного.",
    )

    all_dfs = []
    for chat in chats:
        prog_msg = await q.message.reply_text(f"📡 Парсю `{chat}`...", parse_mode="Markdown")

        async def on_progress(count, _chat=chat, _msg=prog_msg):
            try: await _msg.edit_text(f"📡 `{_chat}`: собрано {count:,}...", parse_mode="Markdown")
            except: pass

        try:
            df = await parse_messages(chat, date_from, date_to, limit, progress_cb=on_progress)
            all_dfs.append(df)
            db_log_parse(user_id, chat, len(df))
            await prog_msg.delete()
        except Exception as e:
            await prog_msg.edit_text(f"❌ Ошибка для `{chat}`: {e}", parse_mode="Markdown")

    if not all_dfs:
        await q.message.reply_text("⚠️ Ничего не найдено.")
        return ConversationHandler.END

    final_df = pd.concat(all_dfs, ignore_index=True)
    final_df["№"] = range(1, len(final_df)+1)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    if fmt == "excel":
        data = df_to_excel(final_df)
        fname = f"tgparse_{ts}.xlsx"
    else:
        data = df_to_csv(final_df)
        fname = f"tgparse_{ts}.csv"

    caption = (
        f"✅ *Готово!*\n\n"
        f"📊 Сообщений: *{len(final_df):,}*\n"
        f"👥 Пользователей: *{final_df['Пользователь'].nunique():,}*\n"
        f"📡 Чатов: *{len(chats)}*"
    )
    await q.message.reply_document(
        document=io.BytesIO(data), filename=fname,
        caption=caption, parse_mode="Markdown",
    )
    return ConversationHandler.END

# ─── Расписание ───────────────────────────────────────────────────────────────
async def cmd_schedule(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    plan = get_user_plan(user_id)

    if not plan["scheduler"]:
        await update.message.reply_text(
            "⏰ *Автопарсинг по расписанию*\n\n"
            "Эта функция доступна в тарифах *PRO* и *MAX*.\n\n"
            "Повысь тариф — /plans",
            parse_mode="Markdown",
        )
        return ConversationHandler.END

    await update.message.reply_text(
        "⏰ *Автопарсинг по расписанию*\n\n"
        "Введи @username чата для автоматического парсинга:",
        parse_mode="Markdown",
    )
    return WAIT_SCHEDULE_CHAT

async def got_schedule_chat(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["sched_chat"] = update.message.text.strip()
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("Каждые 6 часов",  callback_data="si_6"),
         InlineKeyboardButton("Каждые 12 часов", callback_data="si_12")],
        [InlineKeyboardButton("Раз в сутки",     callback_data="si_24"),
         InlineKeyboardButton("Раз в 3 дня",     callback_data="si_72")],
    ])
    await update.message.reply_text("🕐 Как часто парсить?", reply_markup=kb)
    return WAIT_SCHEDULE_INTERVAL

async def got_schedule_interval(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    interval = int(q.data.replace("si_", ""))
    chat = ctx.user_data["sched_chat"]
    user_id = q.from_user.id

    db_add_schedule(user_id, chat, interval)
    await q.edit_message_text(
        f"✅ *Расписание добавлено!*\n\n"
        f"📡 Чат: `{chat}`\n"
        f"🕐 Интервал: каждые {interval} ч.\n\n"
        f"Результаты будут приходить сюда автоматически.",
        parse_mode="Markdown",
    )
    return ConversationHandler.END

# ─── Мой аккаунт ─────────────────────────────────────────────────────────────
async def cmd_account(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    u = db_get_user(user_id)
    plan = get_user_plan(user_id)
    until = u.get("plan_until", "—")
    if until and until != "—":
        until = datetime.fromisoformat(until).strftime("%d.%m.%Y")

    schedules = [s for s in db_get_schedules() if s["user_id"] == user_id]

    text = (
        f"📊 *Мой аккаунт*\n\n"
        f"Тариф: *{plan['name']}*\n"
        f"Действует до: *{until}*\n"
        f"Лимит: *{plan['msg_limit']:,} сообщений*\n"
        f"Чатов: *{plan['chat_limit']}*\n"
        f"Всего спарсено: *{u.get('msgs_used', 0):,}*\n\n"
    )
    if schedules:
        text += f"⏰ *Расписаний: {len(schedules)}*\n"
        for s in schedules:
            text += f"  • `{s['chat']}` — каждые {s['interval_h']}ч\n"

    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("💳 Повысить тариф", callback_data="go_plans")
    ]]) if plan["price"] < PLANS["max"]["price"] else None

    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

async def go_plans(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await cmd_plans(update, ctx)

# ─── Админ статистика ─────────────────────────────────────────────────────────
async def cmd_admin(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return
    s = db_stats()
    await update.message.reply_text(
        f"📈 *Статистика TGParse PRO*\n\n"
        f"👥 Пользователей: *{s['users']:,}*\n"
        f"💳 Платных: *{s['paid']:,}*\n"
        f"⭐ Доход (Stars): *{s['revenue']:,}*\n"
        f"🔄 Парсингов: *{s['parses']:,}*\n"
        f"📨 Сообщений: *{s['msgs']:,}*",
        parse_mode="Markdown",
    )

# ─── Автопарсинг (scheduler) ─────────────────────────────────────────────────
async def run_scheduled_parses(app):
    schedules = db_get_schedules(active_only=True)
    now = datetime.now(timezone.utc)

    for s in schedules:
        last = s.get("last_run")
        if last:
            last_dt = datetime.fromisoformat(last).replace(tzinfo=timezone.utc)
            if (now - last_dt).total_seconds() < s["interval_h"] * 3600:
                continue

        plan = get_user_plan(s["user_id"])
        try:
            df = await parse_messages(s["chat"], now - timedelta(hours=s["interval_h"]), now, plan["msg_limit"])
            if df.empty: continue

            db_log_parse(s["user_id"], s["chat"], len(df))
            data = df_to_excel(df)
            ts = now.strftime("%Y%m%d_%H%M")

            await app.bot.send_document(
                chat_id=s["user_id"],
                document=io.BytesIO(data),
                filename=f"auto_{ts}.xlsx",
                caption=f"⏰ *Автопарсинг* `{s['chat']}`\n📊 {len(df):,} сообщений",
                parse_mode="Markdown",
            )

            con = sqlite3.connect(DB)
            con.execute("UPDATE schedules SET last_run=? WHERE id=?", (now.isoformat(), s["id"]))
            con.commit(); con.close()

        except Exception as e:
            log.error(f"Scheduler error for {s['chat']}: {e}")

# ─── Запуск ───────────────────────────────────────────────────────────────────
def main():
    if not BOT_TOKEN:
        print("❌ TG_BOT_TOKEN не задан в .env"); return

    db_init()
    app = Application.builder().token(BOT_TOKEN).build()

    # Диалог парсинга
    parse_conv = ConversationHandler(
        entry_points=[
            CommandHandler("parse", cmd_parse),
            MessageHandler(filters.Regex("^▶ Парсить$"), cmd_parse),
        ],
        states={
            WAIT_CHAT:   [MessageHandler(filters.TEXT & ~filters.COMMAND, got_chat)],
            WAIT_PERIOD: [CallbackQueryHandler(got_period, pattern="^p_")],
            WAIT_LIMIT:  [CallbackQueryHandler(got_limit,  pattern="^l_")],
            WAIT_FORMAT: [CallbackQueryHandler(got_format, pattern="^f_")],
        },
        fallbacks=[CommandHandler("cancel", lambda u,c: ConversationHandler.END)],
    )

    # Диалог расписания
    sched_conv = ConversationHandler(
        entry_points=[
            CommandHandler("schedule", cmd_schedule),
            MessageHandler(filters.Regex("^📅 Расписание$"), cmd_schedule),
        ],
        states={
            WAIT_SCHEDULE_CHAT:     [MessageHandler(filters.TEXT & ~filters.COMMAND, got_schedule_chat)],
            WAIT_SCHEDULE_INTERVAL: [CallbackQueryHandler(got_schedule_interval, pattern="^si_")],
        },
        fallbacks=[CommandHandler("cancel", lambda u,c: ConversationHandler.END)],
    )

    app.add_handler(CommandHandler("start",   cmd_start))
    app.add_handler(CommandHandler("plans",   cmd_plans))
    app.add_handler(CommandHandler("account", cmd_account))
    app.add_handler(CommandHandler("admin",   cmd_admin))
    app.add_handler(parse_conv)
    app.add_handler(sched_conv)
    app.add_handler(CallbackQueryHandler(handle_buy,  pattern="^buy_"))
    app.add_handler(CallbackQueryHandler(go_plans,    pattern="^go_plans$"))
    app.add_handler(PreCheckoutQueryHandler(pre_checkout))
    app.add_handler(MessageHandler(filters.SUCCESSFUL_PAYMENT, payment_success))
    app.add_handler(MessageHandler(filters.Regex("^💳 Тарифы$"),     cmd_plans))
    app.add_handler(MessageHandler(filters.Regex("^📊 Мой аккаунт$"), cmd_account))

    # Планировщик автопарсинга — запускается внутри event loop
    scheduler = AsyncIOScheduler()
    scheduler.add_job(run_scheduled_parses, "interval", minutes=30, args=[app])

    async def on_startup(application):
        scheduler.start()
        print("⏰ Планировщик запущен!")

    app.post_init = on_startup

    print("🚀 TGParse PRO запущен!")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
